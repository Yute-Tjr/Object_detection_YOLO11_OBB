from __future__ import annotations

import csv
import random
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, Optional, Sequence


@dataclass(frozen=True)
class LabelSample:
    image_name: str
    class_name: str
    group: str


def normalize_status(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip().upper()
    if text in {"OK", "NG"}:
        return text
    return None


def parent_group_key(stem: str) -> str:
    match = re.match(r"^(?P<parent>.+)-\d+$", Path(stem).stem)
    return match.group("parent") if match else Path(stem).stem


def rows_to_label_samples(rows: Iterable[Mapping[str, object]], target_column: str) -> List[LabelSample]:
    samples: List[LabelSample] = []
    for row in rows:
        image_value = row.get("images_name")
        if image_value is None:
            continue
        image_name = str(image_value).strip()
        if not image_name:
            continue
        class_name = normalize_status(row.get(target_column))
        if class_name is None:
            continue
        samples.append(
            LabelSample(
                image_name=Path(image_name).stem,
                class_name=class_name,
                group=parent_group_key(image_name),
            )
        )
    return samples


def split_samples_by_group(
    samples: Sequence[LabelSample],
    train_ratio: float,
    seed: int,
) -> Dict[str, List[LabelSample]]:
    if not 0 < train_ratio < 1:
        raise ValueError("train_ratio must be between 0 and 1")
    groups = sorted({sample.group for sample in samples})
    if not groups:
        raise ValueError("no samples to split")
    shuffled = list(groups)
    random.Random(seed).shuffle(shuffled)
    train_count = int(len(shuffled) * train_ratio)
    if len(shuffled) > 1:
        train_count = min(max(train_count, 1), len(shuffled) - 1)
    train_groups = set(shuffled[:train_count])
    return {
        "train": [sample for sample in samples if sample.group in train_groups],
        "test": [sample for sample in samples if sample.group not in train_groups],
    }


def load_sheet_rows(xlsx_path: Path, sheet_name: str) -> List[Dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required to read Excel label files") from exc

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"sheet not found in {xlsx_path}: {sheet_name}")
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = ["" if value is None else str(value).strip() for value in rows[0]]
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
        for row in rows[1:]
    ]


def write_manifest_csv(path: Path, rows: Iterable[Mapping[str, object]]) -> None:
    fieldnames = [
        "image_name",
        "class_name",
        "group",
        "split",
        "source_json",
        "source_image",
        "crop_path",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
